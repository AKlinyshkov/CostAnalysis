import sqlite3
import sys
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import (QMainWindow, QFileDialog, QApplication)


def export2Excel():
    try:
        # Подключение к базе данных SQLite
        with open("DBWork/DBPath.txt", "r") as dbp:
            dbpath = dbp.read()
        conn = sqlite3.connect(dbpath)
        cursor = conn.cursor()

        # Выполнение SQL-запроса для получения данных
        cursor.execute('SELECT DISTINCT product FROM purchases ORDER BY product')
        rows = cursor.fetchall()

        # Создаем новую книгу Excel и получаем активный лист
        workbook = Workbook()
        sheetPur = workbook.active
        sheetPur.title = "Purchases"

        # Записываем заголовки столбцов
        columns = [description[0] for description in cursor.description]
        sheetPur.append(columns)

        # Записываем данные в лист Excel
        for row in rows:
            sheetPur.append(row)


    ######################################################################
        # Выполнение SQL-запроса для получения данных
        cursor.execute('SELECT category, product FROM categories ORDER BY product')
        rows = cursor.fetchall()

        # Создаем новую книгу Excel и получаем активный лист
        sheetCat = workbook.create_sheet(title="Categories")

        # Записываем заголовки столбцов
        columns = [description[0] for description in cursor.description]
        sheetCat.append(columns)

        # Записываем данные в лист Excel
        for row in rows:
            sheetCat.append(row)

    ########################################################################

        class SelectDBPath(QMainWindow):

            def __init__(self):
                super().__init__()
                self.path = QFileDialog.getExistingDirectory(self, "Select Excel destination", ".")


        app = QApplication(sys.argv)
        xDBPath = SelectDBPath()

        path = xDBPath.path + "/expenses.xlsx"

        # Сохраняем книгу Excel
        workbook.save(path)
    except Exception:
        pass
    finally:
    # Закрываем соединение с базой данных
        conn.close()


def import2Sqlite():

    class SelectxDBPath(QMainWindow):

        def __init__(self):
            super().__init__()
            self.path = QFileDialog.getOpenFileName(self, "Select Excel DB destination", ".", 'All Files (*)')

    class SelectDBPath(QMainWindow):

        def __init__(self):
            super().__init__()
            self.path = QFileDialog.getOpenFileName(self, "Select DB destination", ".", 'All Files (*)')


    app = QApplication(sys.argv)
    xDBPath = SelectxDBPath()
    DBPath = SelectDBPath()

    excel_file_path = xDBPath.path[0]
    workbook = load_workbook(excel_file_path)
    sheet = workbook['Expences']

    # Подключение к базе данных SQLite
    db_path = DBPath.path[0]
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Вставка данных в SQLite
    for row in sheet.iter_rows(min_row=2, max_row=90, values_only=True):
        insert_query = f"INSERT INTO purchases (date, product, description, cost, sum) " \
                       f"VALUES (date('{row[1]}'), '{row[2]}', '{row[3]}', '{row[4]}', '{row[5]}')"
        cursor.execute(insert_query)

    # Сохранение изменений и закрытие соединения
    conn.commit()
    conn.close()


export2Excel()
